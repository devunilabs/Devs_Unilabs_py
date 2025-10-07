# backend/file_processor.py

import os
import pandas as pd
from io import StringIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Mapeo de archivos a hojas
mapeo_archivos = {
    "REPORTFAC_BSD": "BASADRE",
    "REPORTFAC_CGH": "CGH",
    "REPORTFAC_CSI": "CSI",
    "REPORTFAC_HUACHO": "HUACHO",
    "REPORTFAC_AVIMEND": "AVIVA MENDIOLA",
    "REPORTFAC_AVIVACOL": "AVIVA COLONIAL",
    "REPORTFACT_BSD": "BASADRE",
    "REPORTFACT_CGH": "CGH",
    "REPORTFACT_CSI": "CSI",
    "REPORTFACT_HUA": "HUACHO",
    "REPORTFACT_AVI_MEND": "AVIVA MENDIOLA",
    "REPORTFACT_AVI_COL": "AVIVA COLONIAL"
}

# Columnas especiales para determinar el rango a formatear como texto
columnas_especiales = {
    'inicio': ["DNI", "Documento", "DOCUMENTO", "dni"],
    'fin': ["descripciÃ³n prueba", "descripción prueba", "DESCRIPCION PRUEBA", "descripcion prueba"]
}

def determinar_hoja(nombre_archivo):
    nombre_upper = nombre_archivo.upper()
    for prefijo, hoja in mapeo_archivos.items():
        if nombre_upper.startswith(prefijo.upper()):
            return hoja
    return None

def determinar_columnas(df):
    columnas = [col.upper() for col in df.columns]
    
    col_ini = None
    col_fin = None
    
    for patron in columnas_especiales['inicio']:
        if patron.upper() in columnas:
            col_ini = df.columns[columnas.index(patron.upper())]
            break
    
    for patron in columnas_especiales['fin']:
        if patron.upper() in columnas:
            col_fin = df.columns[columnas.index(patron.upper())]
            break
    
    return col_ini, col_fin

def limpiar_contenido(contenido):
    """Limpia el contenido del archivo según la rutina proporcionada"""
    # Limpieza básica: quitar \r\n y tabuladores
    contenido_limpio = contenido.replace('\r\n', ' ').replace('\t', '')
    
    # Limpieza adicional: manejar líneas mal formadas
    lineas = [linea.strip() for linea in contenido_limpio.split('\n') if linea.strip()]
    lineas_limpias = []
    
    for linea in lineas:
        partes = linea.split('|')
        if len(partes) > 39:  # Si tiene más campos de los esperados
            lineas_limpias.append('|'.join(partes[:39]))
        else:
            lineas_limpias.append(linea)
    
    return '\n'.join(lineas_limpias)

def procesar_archivos(archivos, config):
    hojas_data = {}
    advertencias = []
    total_lineas_omitidas = 0
    archivos_procesados = 0

    for archivo in archivos:
        nombre = archivo.filename
        nombre_sin_ext = os.path.splitext(nombre)[0]
        print(f"Procesando archivo: {nombre}")

        try:
            try:
                contenido = archivo.read().decode('utf-8')
            except UnicodeDecodeError:
                archivo.seek(0)
                contenido = archivo.read().decode('ISO-8859-1')
                print(f"Error decodificando {nombre} con UTF-8, intentando ISO-8859-1")

            contenido_limpio = limpiar_contenido(contenido)

            hoja_destino = determinar_hoja(nombre_sin_ext)
            if not hoja_destino:
                advertencias.append(f"Archivo {nombre}: No coincide con ningún patrón de hoja conocido - omitido")
                continue

            df = pd.read_csv(
                StringIO(contenido_limpio),
                sep='|',
                dtype=str,
                engine='python',
                on_bad_lines='warn'
            ).fillna("")

            print(f"Archivo {nombre} leído con {len(df)} filas y {len(df.columns)} columnas.")

            # Control de líneas omitidas
            lineas_esperadas = len(contenido_limpio.split('\n'))
            if len(df) < lineas_esperadas - 1:
                omitidas = lineas_esperadas - len(df) - 1
                total_lineas_omitidas += omitidas
                advertencias.append(f"Archivo {nombre}: {omitidas} líneas omitidas por formato incorrecto")

            col_ini, col_fin = determinar_columnas(df)
            if col_ini and col_fin:
                cols = list(df.columns)
                for col in cols[cols.index(col_ini):cols.index(col_fin)+1]:
                    df[col] = df[col].astype(str).str.strip()
            else:
                advertencias.append(f"Archivo {nombre}: No se encontraron las columnas requeridas")

            hojas_data[hoja_destino] = df
            archivos_procesados += 1

        except Exception as e:
            advertencias.append(f"Error procesando {nombre}: {str(e)} - omitido")
            continue

    if archivos_procesados == 0:
        return {"error": "No se pudo procesar ningún archivo", "advertencias": advertencias}

    fecha_hora = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_archivo = f"REPORTFACT_{fecha_hora}.xlsx"
    ruta_final = os.path.join(config['RESULT_FOLDER'], nombre_archivo)

    try:
        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        for hoja in ["BASADRE", "CGH", "CSI", "HUACHO", "AVIVA MENDIOLA", "AVIVA COLONIAL"]:
            if hoja in hojas_data:
                ws = wb.create_sheet(title=hoja)
                for fila in dataframe_to_rows(hojas_data[hoja], index=False, header=True):
                    ws.append(fila)

                col_ini, col_fin = determinar_columnas(hojas_data[hoja])
                if col_ini and col_fin:
                    inicio = hojas_data[hoja].columns.get_loc(col_ini) + 1
                    fin = hojas_data[hoja].columns.get_loc(col_fin) + 1
                    for row in ws.iter_rows(min_row=2, min_col=inicio, max_col=fin):
                        for cell in row:
                            cell.number_format = '@'

        if not wb.sheetnames:
            wb.create_sheet(title="Resumen")

        wb.save(ruta_final)

        print(f"Archivo Excel generado: {ruta_final}")

        return {
            "status": "success",
            "ruta_final": ruta_final,
            "archivos_procesados": archivos_procesados,
            "total_archivos": len(archivos),
            "total_lineas_omitidas": total_lineas_omitidas,
            "advertencias": advertencias
        }

    except Exception as e:
        return {"error": f"Error al generar el archivo: {str(e)}", "advertencias": advertencias}

