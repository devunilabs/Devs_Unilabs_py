import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
import threading
import time
from datetime import datetime

class UnilabsExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("UNILABS PERU - Procesador archivos DASHBOARD")
        self.root.geometry("1000x700")
        
        # Variables de estado
        self.file_paths = []
        self.processing = False
        self.progress = 0
        self.start_time = 0
        self.start_datetime = ""
        self.end_datetime = ""
        
        # Cargar recursos
        self.load_resources()
        
        # Configurar interfaz
        self.setup_ui()
        
        # Estilo optimizado
        self.setup_styles()
    
    def load_resources(self):
        """Carga la imagen de fondo y otros recursos"""
        try:
            # Ruta a tu imagen UNILABS (cambiar por tu ruta real)
            bg_path = "unilabs_bg.jpg"  # o unilabs_bg.png
            
            # Cargar imagen de fondo
            self.bg_image = Image.open(bg_path) 
            self.bg_image = self.bg_image.resize((1000, 700), Image.LANCZOS)
            self.bg_photo = ImageTk.PhotoImage(self.bg_image)
            
            # Cargar logo
            self.logo_image = Image.open("unilabs_logo.jpg").resize((120, 120), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(self.logo_image)
            
        except Exception as e:
            # Fallback si no hay imagen
            self.bg_photo = None
            messagebox.showwarning("Advertencia"
                                 f"No se pudo cargar imágenes: {str(e)}\nUsando fondo simple.")
    
    def setup_ui(self):
        """Configura la interfaz de usuario"""
        # Canvas principal con fondo
        self.canvas = tk.Canvas(self.root, width=1000, height=700)
        self.canvas.pack(fill="both", expand=True)
        
        if self.bg_photo:
            self.canvas.create_image(0, 0, image=self.bg_photo, anchor="nw")
        else:
            self.canvas.configure(bg='#005baa')  # Fondo azul UNILABS
        
        # Marco principal (semitransparente)
        self.main_frame = tk.Frame(self.canvas, bg='white', bd=3, relief=tk.RIDGE)
        self.main_frame.place(relx=0.5, rely=0.5, anchor="center", width=850, height=550)
        
        # Logo UNILABS en la esquina superior izquierda del main_frame
        if hasattr(self, 'logo_photo'):
            logo_label = tk.Label(self.main_frame, image=self.logo_photo, bg='white')
            logo_label.place(x=20, y=20)  # Posicionamiento absoluto en (20,20)
        else:
            logo_label = tk.Label(self.main_frame, text="UNILABS PERU", font=('Arial', 14, 'bold'), 
                                fg='#005baa', bg='white')
            logo_label.place(x=20, y=20)
        
        # Título (ajustado para no solapar con el logo)
        tk.Label(self.main_frame, text="Procesador archivos DASHBOARD", 
                font=('Arial', 16), fg='#333333', bg='white').pack(pady=(80, 5))
        
        # Controles principales
        self.setup_controls()
        
        # Frame para información de tiempo
        time_frame = tk.Frame(self.main_frame, bg='white')
        time_frame.pack(pady=5)
        
        # Etiquetas para mostrar fechas
        self.start_time_label = tk.Label(time_frame, text="Inicio: -", font=('Arial', 9), 
                                       fg='#333333', bg='white')
        self.start_time_label.pack(side=tk.LEFT, padx=10)
        
        self.end_time_label = tk.Label(time_frame, text="Fin: -", font=('Arial', 9), 
                                     fg='#333333', bg='white')
        self.end_time_label.pack(side=tk.LEFT, padx=10)
        
        # Estadísticas
        self.stats_label = tk.Label(self.main_frame, text="", font=('Arial', 9), 
                                   fg='#666666', bg='white')
        self.stats_label.pack(pady=5)
    
    def setup_styles(self):
        """Configura estilos visuales"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Estilo para botones
        style.configure('Unilabs.TButton', 
                      foreground='white', background='#005baa',
                      font=('Arial', 11, 'bold'),
                      padding=8)
        
        # Estilo para barra de progreso
        style.configure('Unilabs.Horizontal.TProgressbar',
                      background='#005baa',
                      troughcolor='#e0e0e0',
                      thickness=20)
    
    def setup_controls(self):
        """Configura los controles de la interfaz"""
        # Frame para botones
        btn_frame = tk.Frame(self.main_frame, bg='white')
        btn_frame.pack(pady=10)
        
        # Botón para cargar archivos
        self.load_btn = ttk.Button(btn_frame, text="SELECCIONA TUS ARCHIVOS DASH", 
                                 command=self.load_files, style='Unilabs.TButton')
        self.load_btn.pack(side=tk.LEFT, padx=5)
        
        # Botón para procesar
        self.process_btn = ttk.Button(btn_frame, text="PROCESAR", 
                                    command=self.start_processing, 
                                    style='Unilabs.TButton', state=tk.DISABLED)
        self.process_btn.pack(side=tk.LEFT, padx=5)
        
        # Lista de archivos con scroll
        list_frame = tk.Frame(self.main_frame, bg='white')
        list_frame.pack(pady=10)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.file_listbox = tk.Listbox(list_frame, width=90, height=12,
                                     yscrollcommand=scrollbar.set,
                                     selectbackground='#005baa',
                                     selectforeground='white',
                                     font=('Arial', 10))
        self.file_listbox.pack(side=tk.LEFT)
        scrollbar.config(command=self.file_listbox.yview)
        
        # Barra de progreso
        progress_frame = tk.Frame(self.main_frame, bg='white')
        progress_frame.pack(pady=10)
        
        self.progress_label = tk.Label(progress_frame, text="0% completado", 
                                     font=('Arial', 10, 'bold'), bg='white')
        self.progress_label.pack()
        
        self.progress_bar = ttk.Progressbar(progress_frame, length=600,
                                          style='Unilabs.Horizontal.TProgressbar')
        self.progress_bar.pack(pady=5)
        
        # Etiqueta de estado
        self.status_label = tk.Label(self.main_frame, text="", 
                                    font=('Arial', 10), bg='white')
        self.status_label.pack(pady=5)
    
    def load_files(self):
        """Carga los archivos Excel a procesar"""
        files = filedialog.askopenfilenames(
            title="Seleccionar archivos Excel",
            filetypes=(("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")))
        
        if files:
            self.file_paths = list(files)[:7]  # Limitar a 7 archivos
            self.file_listbox.delete(0, tk.END)
            
            for file_path in self.file_paths:
                self.file_listbox.insert(tk.END, os.path.basename(file_path))
            
            self.process_btn.config(state=tk.NORMAL)
            self.status_label.config(text=f"{len(self.file_paths)} archivos listos", fg='blue')
            self.stats_label.config(text="")
            self.start_time_label.config(text="Inicio: -")
            self.end_time_label.config(text="Fin: -")
    
    def start_processing(self):
        """Inicia el procesamiento de los archivos"""
        if not self.file_paths:
            messagebox.showwarning("Advertencia", "No hay archivos para procesar")
            return
        
        self.processing = True
        self.progress = 0
        self.start_time = time.time()
        self.start_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.update_progress(0)
        
        # Actualizar etiqueta de inicio
        self.start_time_label.config(text=f"Inicio: {self.start_datetime}")
        self.end_time_label.config(text="Fin: -")
        
        # Deshabilitar controles durante el procesamiento
        self.load_btn.config(state=tk.DISABLED)
        self.process_btn.config(state=tk.DISABLED)
        
        # Usar hilo para no bloquear la interfaz
        threading.Thread(target=self.process_files, daemon=True).start()
    
    def process_files(self):
        """Procesa los archivos en segundo plano"""
        try:
            total_files = len(self.file_paths)
            
            for i, file_path in enumerate(self.file_paths):
                if not self.processing:
                    break
                
                # Procesar el archivo actual
                self.process_excel_file(file_path)
                
                # Actualizar progreso
                progress = ((i + 1) / total_files) * 100
                self.update_progress(progress)
                
                # Actualizar estado
                self.root.after(10, lambda f=os.path.basename(file_path), 
                              idx=i+1, tot=total_files: 
                              self.status_label.config(
                                  text=f"Procesando: {f} ({idx}/{tot})"))
            
            if self.processing:
                elapsed = time.time() - self.start_time
                self.end_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                
                self.stats_label.config(
                    text=f"Procesado {total_files} archivos en {elapsed:.2f} segundos")
                
                self.root.after(10, lambda: [
                    self.status_label.config(text="¡Proceso completado!", fg='green'),
                    self.end_time_label.config(text=f"Fin: {self.end_datetime}"),
                    messagebox.showinfo("Completado", 
                                     f"Procesado {total_files} archivos en {elapsed:.2f} segundos\n"
                                     f"Inicio: {self.start_datetime}\n"
                                     f"Fin: {self.end_datetime}"),
                    self.load_btn.config(state=tk.NORMAL)
                ])
        
        except Exception as e:
            self.end_datetime = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.root.after(10, lambda: [
                self.status_label.config(text=f"Error: {str(e)}", fg='red'),
                self.end_time_label.config(text=f"Fin: {self.end_datetime}"),
                messagebox.showerror("Error", f"Ocurrió un error: {str(e)}\n"
                                           f"Inicio: {self.start_datetime}\n"
                                           f"Fin: {self.end_datetime}"),
                self.load_btn.config(state=tk.NORMAL),
                self.process_btn.config(state=tk.NORMAL if self.file_paths else tk.DISABLED)
            ])
        
        finally:
            self.processing = False
    
    def process_excel_file(self, file_path):
        """Procesa un archivo Excel individual"""
        try:
            # Usar modo read_only para carga más rápida
            wb = load_workbook(filename=file_path, read_only=False)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Insertar columna LLAVE al principio
                ws.insert_cols(1)
                ws['A1'] = "LLAVE"
                
                # Copiar estilo de la celda B1 a A1
                if 'B1' in ws:
                    ws['A1']._style = ws['B1']._style
                
                # Procesamiento optimizado por lotes
                max_row = ws.max_row
                values = []
                
                # Pre-calcular referencias de columnas
                col_b = get_column_letter(2)
                col_bc = get_column_letter(55)
                
                # Procesar todas las filas
                for row in range(2, max_row + 1):
                    # Obtener valores directamente (sin fórmulas)
                    val_b = ws[f'{col_b}{row}'].value or ""
                    val_bc = ws[f'{col_bc}{row}'].value or ""
                    
                    # Concatenar y guardar
                    ws[f'A{row}'] = f"{val_b}{val_bc}"
            
            # Guardar cambios
            wb.save(file_path)
            wb.close()
        
        except Exception as e:
            raise Exception(f"Error procesando {os.path.basename(file_path)}: {str(e)}")
    
    def update_progress(self, value):
        """Actualiza la barra de progreso"""
        self.progress = value
        self.progress_bar['value'] = value
        self.progress_label.config(text=f"{int(value)}% completado")
        self.root.update_idletasks()
    
    def on_closing(self):
        """Maneja el cierre de la aplicación"""
        self.processing = False
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    
    # Configuración para mejor rendimiento
    root.tk.call('tk', 'scaling', 1.5)  # Mejorar escalado en pantallas HD
    
    app = UnilabsExcelProcessor(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()
    root = tk.Tk()
    
    # Configuración para mejor rendimiento
    root.tk.call('tk', 'scaling', 1.5)  # Mejorar escalado en pantallas HD
    
    app = UnilabsExcelProcessor(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()