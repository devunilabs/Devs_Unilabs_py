import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm
from threading import Thread
import shutil

# Ruta de Descargas del usuario
descargas_path = os.path.join(os.path.expanduser("~"), "Downloads")

def procesar_archivos(rutas, progreso_label):
    total = len(rutas)
    
    for i, archivo in enumerate(rutas):
        nombre = os.path.basename(archivo)
        progreso_label.config(text=f"Procesando {nombre} ({i+1}/{total})...")
        root.update()

        # Cargar workbook y hoja activa
        wb = load_workbook(archivo)
        ws: Worksheet = wb.active
        
        max_fila = ws.max_row

        # Insertar nueva columna A con cabecera "LLAVE"
        ws.insert_cols(1)
        ws['A1'] = 'LLAVE'

        for fila in tqdm(range(2, max_fila + 1), desc=f"Procesando {nombre}", leave=False):
            val_b = ws[f'B{fila}'].value or ''
            val_bc = ws[f'BC{fila}'].value or ''
            ws[f'A{fila}'] = f"{val_b}{val_bc}"

        # Guardar en la carpeta Descargas
        nuevo_nombre = os.path.join(descargas_path, f"procesado_{nombre}")
        wb.save(nuevo_nombre)
        wb.close()
    
    progreso_label.config(text="¡Todos los archivos fueron procesados con éxito!")
    messagebox.showinfo("Finalizado", "Archivos procesados correctamente y guardados en Descargas.")

def seleccionar_archivos():
    rutas = filedialog.askopenfilenames(title="Selecciona los archivos Excel", filetypes=[("Archivos Excel", "*.xlsx")])
    if rutas:
        Thread(target=procesar_archivos, args=(rutas, progreso_label)).start()

# Interfaz gráfica
root = tk.Tk()
root.title("Procesador de Archivos Excel")
root.geometry("450x180")
root.resizable(False, False)

frame = tk.Frame(root, padx=20, pady=20)
frame.pack()

tk.Label(frame, text="Procesador de Excel con columna LLAVE", font=("Arial", 14)).pack(pady=10)

boton = tk.Button(frame, text="Seleccionar archivos", font=("Arial", 12), command=seleccionar_archivos)
boton.pack(pady=10)

progreso_label = tk.Label(frame, text="", font=("Arial", 10))
progreso_label.pack()

root.mainloop()
