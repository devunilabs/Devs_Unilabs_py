import os
import pandas as pd
from flask import Flask, request, render_template_string, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import StringIO
from datetime import datetime

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

RESULT_FOLDER = os.path.join(os.path.expanduser("~"), 'Downloads')

# Mapeo de archivos a hojas
mapeo_archivos = {
    "REPORTFAC_BSD": "BASADRE",
    "REPORTFAC_CGH": "CGH",
    "REPORTFAC_CSI": "CSI",
    "REPORTFAC_HUACHO": "HUACHO",
    "REPORTFAC_AVIMEND": "AVIVA MENDIOLA",
    "REPORTFAC_AVIVACOL": "AVIVA COLONIAL",
    "REPORTFACT_BSD": "BASADRE",
    "REPORTFACT_CGH": "CGH",
    "REPORTFACT_CSI": "CSI",
    "REPORTFACT_HUA": "HUACHO",
    "REPORTFACT_AVI_MEND": "AVIVA MENDIOLA",
    "REPORTFACT_AVI_COL": "AVIVA COLONIAL"
}

columnas_especiales = {
    'inicio': ["DNI", "Documento", "DOCUMENTO", "dni"],
    'fin': ["descripciÃ³n prueba", "descripción prueba", "DESCRIPCION PRUEBA", "descripcion prueba"]
}

# HTML Template completo
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Procesador REPORTFACT</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css">
    <style>
        .card {
            border-radius: 15px;
            box-shadow: 0 6px 10px rgba(0,0,0,0.08);
        }
        .progress {
            height: 25px;
            border-radius: 12px;
        }
        .btn-primary {
            background-color: #4e73df;
            border: none;
            padding: 10px 25px;
            font-weight: 600;
        }
        .status-card {
            transition: all 0.3s ease;
        }
        .status-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 10px 20px rgba(0,0,0,0.1);
        }
        .file-input-label {
            display: block;
            padding: 15px;
            border: 2px dashed #ccc;
            border-radius: 10px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s;
        }
        .file-input-label:hover {
            border-color: #4e73df;
            background-color: #f8f9fa;
        }
    </style>
</head>
<body class="bg-light">
    <div class="container py-5">
        <div class="row justify-content-center">
            <div class="col-lg-10">
                <div class="card shadow-sm p-4 mb-5 bg-white rounded">
                    <div class="text-center mb-4">
                        <h2 class="fw-bold text-primary">
                            <i class="bi bi-file-earmark-spreadsheet"></i> Procesador REPORTFACT
                        </h2>
                        <p class="text-muted">Versión 1.0 - Procesamiento Mensual y Semanal</p>
                        <div class="alert alert-info">
                            <i class="bi bi-info-circle"></i> Seleccione los archivos a procesar (REPORTFACT_*.csv)
                        </div>
                    </div>
                    
                    <div class="mb-4">
                        <form id="uploadForm" enctype="multipart/form-data">
                            <input type="file" id="fileInput" name="archivos" multiple required style="display: none;">
                            <label for="fileInput" class="file-input-label mb-3">
                                <i class="bi bi-cloud-arrow-up" style="font-size: 2rem;"></i>
                                <div class="mt-2">Haga clic para seleccionar archivos</div>
                                <small class="text-muted" id="fileCount">No se han seleccionado archivos</small>
                            </label>
                            
                            <div class="d-grid gap-2">
                                <button type="submit" class="btn btn-primary btn-lg" id="processBtn">
                                    <i class="bi bi-gear"></i> Procesar Archivos
                                </button>
                            </div>
                        </form>
                    </div>
                    
                    <div id="progressSection" style="display: none;">
                        <div class="d-flex justify-content-between mb-2">
                            <span class="fw-bold">Progreso general:</span>
                            <span id="progressText">0%</span>
                        </div>
                        <div class="progress mb-4">
                            <div id="progressBar" class="progress-bar progress-bar-striped progress-bar-animated" 
                                 role="progressbar" style="width: 0%"></div>
                        </div>
                        
                        <h5 class="mb-3"><i class="bi bi-list-check"></i> Progreso por archivo:</h5>
                        <div class="row" id="statusCards">
                            <!-- Aquí se agregarán las tarjetas de estado dinámicamente -->
                        </div>
                    </div>
                    
                    <div id="resultSection" style="display: none;">
                        <div class="alert alert-success" role="alert">
                            <h4 class="alert-heading">
                                <i class="bi bi-check-circle"></i> Procesamiento completado!
                            </h4>
                            <div id="resultContent"></div>
                            <hr>
                            <div class="d-grid gap-2">
                                <button class="btn btn-outline-primary" onclick="location.reload()">
                                    <i class="bi bi-arrow-repeat"></i> Procesar más archivos
                                </button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        // Mostrar cantidad de archivos seleccionados
        document.getElementById('fileInput').addEventListener('change', function(e) {
            const files = e.target.files;
            const fileCount = document.getElementById('fileCount');
            
            if (files.length > 0) {
                fileCount.textContent = `${files.length} archivo(s) seleccionado(s)`;
                fileCount.className = 'text-success';
            } else {
                fileCount.textContent = 'No se han seleccionado archivos';
                fileCount.className = 'text-muted';
            }
        });
        
        // Manejar envío del formulario
        document.getElementById('uploadForm').addEventListener('submit', function(e) {
            e.preventDefault();
            
            const formData = new FormData();
            const files = document.getElementById('fileInput').files;
            
            for (let i = 0; i < files.length; i++) {
                formData.append('archivos', files[i]);
            }
            
            // Mostrar sección de progreso
            document.getElementById('progressSection').style.display = 'block';
            document.getElementById('processBtn').disabled = true;
            
            // Configurar tarjetas de estado
            const statusCards = document.getElementById('statusCards');
            statusCards.innerHTML = '';
            
            for (let i = 0; i < files.length; i++) {
                statusCards.innerHTML += `
                    <div class="col-md-6 mb-3">
                        <div class="card status-card h-100">
                            <div class="card-body">
                                <h5 class="card-title d-flex justify-content-between">
                                    <span class="text-truncate">${files[i].name}</span>
                                    <span id="statusIcon${i}" class="text-muted">
                                        <i class="bi bi-hourglass"></i>
                                    </span>
                                </h5>
                                <div class="progress" style="height: 8px;">
                                    <div id="fileProgress${i}" class="progress-bar" role="progressbar" style="width: 0%"></div>
                                </div>
                                <small id="statusText${i}" class="text-muted">En cola</small>
                            </div>
                        </div>
                    </div>
                `;
            }
            
            // Enviar archivos al servidor
            const xhr = new XMLHttpRequest();
            xhr.open('POST', '/procesar', true);
            
            xhr.upload.onprogress = function(e) {
                if (e.lengthComputable) {
                    const percent = Math.round((e.loaded / e.total) * 100);
                    document.getElementById('progressBar').style.width = percent + '%';
                    document.getElementById('progressText').textContent = percent + '%';
                }
            };
            
            xhr.onload = function() {
                if (xhr.status === 200) {
                    const response = JSON.parse(xhr.responseText);
                    document.getElementById('resultContent').innerHTML = `
                        <p><i class="bi bi-file-earmark-excel"></i> <strong>Archivo generado:</strong> ${response.ruta_final}</p>
                        <p><i class="bi bi-check2-circle"></i> <strong>Archivos procesados:</strong> ${response.archivos_procesados} de ${files.length}</p>
                        <p><i class="bi bi-exclamation-triangle"></i> <strong>Líneas omitidas:</strong> ${response.total_lineas_omitidas}</p>
                        ${response.advertencias.length > 0 ? 
                            `<div class="mt-3 alert alert-warning">
                                <h5><i class="bi bi-exclamation-octagon"></i> Advertencias:</h5>
                                <ul class="mb-0">${response.advertencias.map(adv => `<li>${adv}</li>`).join('')}</ul>
                            </div>` : 
                            '<div class="mt-3 alert alert-success"><i class="bi bi-check-all"></i> Todos los archivos se procesaron correctamente</div>'}
                    `;
                    
                    document.getElementById('progressSection').style.display = 'none';
                    document.getElementById('resultSection').style.display = 'block';
                } else {
                    alert('Error al procesar los archivos: ' + xhr.responseText);
                    document.getElementById('processBtn').disabled = false;
                }
            };
            
            xhr.onerror = function() {
                alert('Error de conexión al servidor');
                document.getElementById('processBtn').disabled = false;
            };
            
            xhr.send(formData);
            
            // Simular progreso individual de archivos
            for (let i = 0; i < files.length; i++) {
                simulateFileProgress(i, files.length);
            }
        });
        
        function simulateFileProgress(index, totalFiles) {
            let progress = 0;
            const interval = setInterval(() => {
                progress += Math.random() * 10;
                if (progress >= 100) {
                    progress = 100;
                    clearInterval(interval);
                    document.getElementById(`statusIcon${index}`).innerHTML = '<i class="bi bi-check-circle text-success"></i>';
                    document.getElementById(`statusText${index}`).textContent = 'Completado';
                    document.getElementById(`statusText${index}`).className = 'text-success';
                }
                document.getElementById(`fileProgress${index}`).style.width = progress + '%';
                document.getElementById(`statusText${index}`).textContent = `${Math.round(progress)}% procesado`;
            }, 300 + (index * 100));
        }
    </script>
</body>
</html>
"""

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

def determinar_hoja(nombre_archivo):
    nombre_upper = nombre_archivo.upper()
    for prefijo, hoja in mapeo_archivos.items():
        if nombre_upper.startswith(prefijo.upper()):
            return hoja
    return None

def determinar_columnas(df):
    columnas = [col.upper() for col in df.columns]
    
    for patron in columnas_especiales['inicio']:
        if patron.upper() in columnas:
            col_ini = df.columns[columnas.index(patron.upper())]
            break
    else:
        col_ini = None
    
    for patron in columnas_especiales['fin']:
        if patron.upper() in columnas:
            col_fin = df.columns[columnas.index(patron.upper())]
            break
    else:
        col_fin = None
    
    return col_ini, col_fin

def limpiar_contenido(contenido):
    """Limpia el contenido del archivo según la rutina proporcionada"""
    # Limpieza básica: quitar \r\n y tabuladores
    contenido_limpio = contenido.replace('\r\n', ' ').replace('\t', '')
    
    # Limpieza adicional: manejar líneas mal formadas
    lineas = [linea.strip() for linea in contenido_limpio.split('\n') if linea.strip()]
    lineas_limpias = []
    
    for linea in lineas:
        partes = linea.split('|')
        if len(partes) > 39:  # Si tiene más campos de los esperados
            lineas_limpias.append('|'.join(partes[:39]))
        else:
            lineas_limpias.append(linea)
    
    return '\n'.join(lineas_limpias)

@app.route('/procesar', methods=['POST'])
def procesar():
    archivos = request.files.getlist("archivos")
    hojas_data = {}
    advertencias = []
    total_lineas_omitidas = 0
    archivos_procesados = 0

    for archivo in archivos:
        nombre = secure_filename(archivo.filename)
        nombre_sin_ext = os.path.splitext(nombre)[0]
        
        try:
            contenido = archivo.read().decode('utf-8', errors='ignore')
            contenido_limpio = limpiar_contenido(contenido)

            hoja_destino = determinar_hoja(nombre_sin_ext)
            if not hoja_destino:
                advertencias.append(f"Archivo {nombre}: No coincide con ningún patrón de hoja conocido - omitido")
                continue

            try:
                df = pd.read_csv(
                    StringIO(contenido_limpio),
                    sep='|',
                    dtype=str,
                    engine='python',
                    on_bad_lines='warn'
                ).fillna("")
                
                lineas_esperadas = len(contenido_limpio.split('\n'))
                if len(df) < lineas_esperadas - 1:
                    omitidas = lineas_esperadas - len(df) - 1
                    total_lineas_omitidas += omitidas
                    advertencias.append(f"Archivo {nombre}: {omitidas} líneas omitidas por formato incorrecto")
            except Exception as e:
                advertencias.append(f"Error al leer {nombre}: {str(e)} - omitido")
                continue

            col_ini, col_fin = determinar_columnas(df)
            if col_ini and col_fin:
                cols = list(df.columns)
                for col in cols[cols.index(col_ini):cols.index(col_fin)+1]:
                    df[col] = df[col].astype(str).str.strip()
            else:
                advertencias.append(f"Archivo {nombre}: No se encontraron las columnas requeridas")

            hojas_data[hoja_destino] = df
            archivos_procesados += 1

        except Exception as e:
            advertencias.append(f"Error procesando {nombre}: {str(e)} - omitido")
            continue

    if archivos_procesados == 0:
        return jsonify({
            "error": "No se pudo procesar ningún archivo",
            "advertencias": advertencias
        }), 400

    fecha_hora = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_archivo = f"REPORTFACT_{fecha_hora}.xlsx"
    ruta_final = os.path.join(RESULT_FOLDER, nombre_archivo)

    try:
        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        for hoja in ["BASADRE", "CGH", "CSI", "HUACHO", "AVIVA MENDIOLA", "AVIVA COLONIAL"]:
            if hoja in hojas_data:
                ws = wb.create_sheet(title=hoja)
                for fila in dataframe_to_rows(hojas_data[hoja], index=False, header=True):
                    ws.append(fila)
                
                col_ini, col_fin = determinar_columnas(hojas_data[hoja])
                if col_ini and col_fin:
                    inicio = hojas_data[hoja].columns.get_loc(col_ini) + 1
                    fin = hojas_data[hoja].columns.get_loc(col_fin) + 1
                    for row in ws.iter_rows(min_row=2, min_col=inicio, max_col=fin):
                        for cell in row:
                            cell.number_format = '@'

        if not wb.sheetnames:
            wb.create_sheet(title="Resumen")

        wb.save(ruta_final)

        return jsonify({
            "status": "success",
            "ruta_final": ruta_final,
            "archivos_procesados": archivos_procesados,
            "total_archivos": len(archivos),
            "total_lineas_omitidas": total_lineas_omitidas,
            "advertencias": advertencias
        })

    except Exception as e:
        return jsonify({
            "error": f"Error al generar el archivo: {str(e)}",
            "advertencias": advertencias
        }), 500

if __name__ == '__main__':
    app.run(debug=True)