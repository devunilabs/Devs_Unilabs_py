#version 1.0 para procesar los archivos reportfact Mensual y Semanal
import os
import pandas as pd
from flask import Flask, request, render_template
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import StringIO
from datetime import datetime

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

RESULT_FOLDER = os.path.join(os.path.expanduser("~"), 'Downloads')

# Mapeo actualizado con los patrones de nombre
mapeo_archivos = {
    "REPORTFAC_BSD": "BASADRE",
    "REPORTFAC_CGH": "CGH",
    "REPORTFAC_CSI": "CSI",
    "REPORTFAC_HUACHO": "HUACHO",
    "REPORTFAC_AVIMEND": "AVIVA MENDIOLA",
    "REPORTFAC_AVIVACOL": "AVIVA COLONIAL",
    "REPORTFACT_BSD": "BASADRE",
    "REPORTFACT_CGH": "CGH",
    "REPORTFACT_CSI": "CSI",
    "REPORTFACT_HUA": "HUACHO",
    "REPORTFACT_AVI_MEND": "AVIVA MENDIOLA",
    "REPORTFACT_AVI_COL": "AVIVA COLONIAL"
}

# Variantes de nombres de columnas
columnas_especiales = {
    'inicio': ["DNI", "Documento", "DOCUMENTO", "dni"],
    'fin': ["descripciÃ³n prueba", "descripción prueba", "DESCRIPCION PRUEBA", "descripcion prueba"]
}

@app.route('/')
def index():
    return render_template('index.html')

def determinar_hoja(nombre_archivo):
    """Determina la hoja basada en el patrón del nombre del archivo"""
    nombre_upper = nombre_archivo.upper()
    for prefijo, hoja in mapeo_archivos.items():
        if nombre_upper.startswith(prefijo.upper()):
            return hoja
    return None

def determinar_columnas(df):
    """Identifica las columnas de inicio y fin con tolerancia a variaciones"""
    columnas = [col.upper() for col in df.columns]
    
    # Buscar columna inicial
    for patron in columnas_especiales['inicio']:
        if patron.upper() in columnas:
            col_ini = df.columns[columnas.index(patron.upper())]
            break
    else:
        col_ini = None
    
    # Buscar columna final
    for patron in columnas_especiales['fin']:
        if patron.upper() in columnas:
            col_fin = df.columns[columnas.index(patron.upper())]
            break
    else:
        col_fin = None
    
    return col_ini, col_fin

def limpiar_contenido(contenido):
    """Limpia el contenido del archivo según la rutina proporcionada"""
    # Limpieza básica: quitar \r\n y tabuladores
    contenido_limpio = contenido.replace('\r\n', ' ').replace('\t', '')
    
    # Limpieza adicional: manejar líneas mal formadas
    lineas = [linea.strip() for linea in contenido_limpio.split('\n') if linea.strip()]
    lineas_limpias = []
    
    for linea in lineas:
        partes = linea.split('|')
        if len(partes) > 39:  # Si tiene más campos de los esperados
            lineas_limpias.append('|'.join(partes[:39]))
        else:
            lineas_limpias.append(linea)
    
    return '\n'.join(lineas_limpias)

@app.route('/procesar', methods=['POST'])
def procesar():
    archivos = request.files.getlist("archivos")
    hojas_data = {}
    advertencias = []
    total_lineas_omitidas = 0
    archivos_procesados = 0

    for archivo in archivos:
        nombre = secure_filename(archivo.filename)
        nombre_sin_ext = os.path.splitext(nombre)[0]
        
        try:
            # Leer y limpiar contenido (según la rutina proporcionada)
            contenido_original = archivo.read().decode('utf-8', errors='ignore')
            contenido_limpio = limpiar_contenido(contenido_original)

            # Detectar hoja destino (versión mejorada)
            hoja_destino = determinar_hoja(nombre_sin_ext)
            if not hoja_destino:
                advertencias.append(f"Archivo {nombre}: No coincide con ningún patrón de hoja conocido - omitido")
                continue

            # Leer CSV con manejo de errores
            try:
                df = pd.read_csv(
                    StringIO(contenido_limpio),
                    sep='|',
                    dtype=str,
                    engine='python',
                    on_bad_lines='warn'
                )
                df = df.fillna("")
                
                # Verificar líneas omitidas
                lineas_esperadas = len(contenido_limpio.split('\n'))
                if len(df) < lineas_esperadas - 1:  # Restar 1 por el encabezado
                    omitidas = lineas_esperadas - len(df) - 1
                    total_lineas_omitidas += omitidas
                    advertencias.append(f"Archivo {nombre}: {omitidas} líneas omitidas por formato incorrecto")
            except Exception as e:
                advertencias.append(f"Error al leer {nombre}: {str(e)} - omitido")
                continue

            # Convertir columnas específicas a texto (versión mejorada)
            col_ini, col_fin = determinar_columnas(df)
            if col_ini and col_fin:
                columnas = list(df.columns)
                i_ini = columnas.index(col_ini)
                i_fin = columnas.index(col_fin)
                cols_texto = columnas[i_ini:i_fin+1]
                
                for col in cols_texto:
                    df[col] = df[col].astype(str).str.strip()
            else:
                advertencias.append(f"Archivo {nombre}: No se encontraron las columnas requeridas (DNI/Documento y descripción prueba)")

            hojas_data[hoja_destino] = df
            archivos_procesados += 1

        except Exception as e:
            advertencias.append(f"Error procesando {nombre}: {str(e)} - omitido")
            continue

    # Manejar caso cuando no se procesa ningún archivo
    if archivos_procesados == 0:
        return f"""
        <h2>❌ Error en el procesamiento</h2>
        <p>No se pudo procesar ningún archivo. Razones:</p>
        <ul>
        {"".join(f"<li>{adv}</li>" for adv in advertencias)}
        </ul>
        <p>Verifique que los nombres de archivo coincidan con estos patrones:</p>
        <ul>
        {"".join(f"<li>{prefijo}_* (para {hoja})</li>" for prefijo, hoja in mapeo_archivos.items())}
        </ul>
        <a href="/">Volver</a>
        """

    # Generar archivo Excel
    fecha_hora = datetime.now().strftime("%Y%m%d_%H")
    nombre_archivo = f"REPORTFACT_{fecha_hora}.xlsx"
    ruta_final = os.path.join(RESULT_FOLDER, nombre_archivo)

    try:
        wb = Workbook()
        if len(wb.sheetnames) > 0:
            wb.remove(wb.active)

        # Orden de hojas en el Excel
        orden_hojas = [
            "BASADRE", "CGH", "CSI", "HUACHO",
            "AVIVA MENDIOLA", "AVIVA COLONIAL"
        ]

        for hoja in orden_hojas:
            if hoja in hojas_data:
                df = hojas_data[hoja]
                ws = wb.create_sheet(title=hoja)
                
                for fila in dataframe_to_rows(df, index=False, header=True):
                    ws.append(fila)
                
                # Aplicar formato texto si se encontraron columnas
                col_ini, col_fin = determinar_columnas(df)
                if col_ini and col_fin:
                    inicio_idx = df.columns.get_loc(col_ini) + 1
                    fin_idx = df.columns.get_loc(col_fin) + 1
                    
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                         min_col=inicio_idx, max_col=fin_idx):
                        for cell in row:
                            cell.number_format = '@'

        # Asegurar que hay al menos una hoja
        if len(wb.sheetnames) == 0:
            wb.create_sheet(title="Datos")

        wb.save(ruta_final)

        # Mensaje de resumen
        mensaje = f"""
        <h2>✅ Procesamiento completado</h2>
        <p>Archivo generado: <b>{ruta_final}</b></p>
        <p>Archivos procesados: {archivos_procesados} de {len(archivos)}</p>
        <p>Total líneas omitidas: {total_lineas_omitidas}</p>
        """
        
        if advertencias:
            mensaje += "<h3>⚠️ Advertencias:</h3><ul>"
            mensaje += "".join(f"<li>{adv}</li>" for adv in advertencias)
            mensaje += "</ul>"
        
        mensaje += '<a href="/">Volver</a>'
        return mensaje

    except Exception as e:
        return f"""
        <h2>❌ Error al generar el archivo</h2>
        <p><b>Error:</b> {str(e)}</p>
        <h3>Advertencias previas:</h3>
        <ul>{"".join(f"<li>{adv}</li>" for adv in advertencias)}</ul>
        <a href="/">Volver</a>
        """

if __name__ == '__main__':
    app.run(debug=True)
